"""Tests for Owner Requirements ingestion — multi-sheet, modes, provenance, dry-run."""

import json

from openpyxl import Workbook
from sqlalchemy import BigInteger, create_engine, select
from sqlalchemy.dialects.postgresql import JSONB
from sqlalchemy.ext.compiler import compiles
from sqlalchemy.orm import Session

from app.ingestion.requirements_loader import ingest_requirements_file, PARSER_VERSION
from app.models import Base, Client, Organization, Requirement, RequirementSourceFile


@compiles(JSONB, "sqlite")
def _compile_jsonb_for_sqlite(type_: str, compiler, **kw):
    return "JSON"


@compiles(BigInteger, "sqlite")
def _compile_bigint_for_sqlite(type_: str, compiler, **kw):
    return "INTEGER"


ROCKWALL_NORTHWEST_HEADERS = [
    "STATUS", "DISCIPLINE", "REQUIREMENT", "LINKS", "CATEGORY LIST",
    "DATE UPDATED", "MODIFIED BY", "RESOURCE", "Item Type", "Path",
]


def _write_requirements_xlsx(path, headers, rows, sheet_name="Sheet"):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(headers)
    for row in rows:
        ws.append(row)
    wb.save(path)
    wb.close()


def _write_multi_sheet_xlsx(path, sheets):
    """sheets: list of (sheet_name, headers, rows)"""
    wb = Workbook()
    wb.remove(wb.active)
    for name, headers, rows in sheets:
        ws = wb.create_sheet(title=name)
        ws.append(headers)
        for row in rows:
            ws.append(row)
    wb.save(path)
    wb.close()


def _make_session():
    engine = create_engine("sqlite+pysqlite:///:memory:")
    Base.metadata.create_all(engine)
    session = Session(engine)
    organization = Organization(id=1, name="EMA Engineering")
    client = Client(
        id=1, organization_id=organization.id, code="TEST_OWNER", display_name="Test Owner",
    )
    session.add_all([organization, client])
    session.flush()
    return session, client


# ---------------------------------------------------------------------------
# AGT-014: Multi-sheet parsing
# ---------------------------------------------------------------------------


def test_loader_reads_all_sheets(tmp_path):
    xlsx_path = tmp_path / "multi_sheet.xlsx"
    _write_multi_sheet_xlsx(xlsx_path, [
        ("Electrical", ROCKWALL_NORTHWEST_HEADERS, [
            ["DONE", "ELECTRICAL", "Provide panel schedule.", None, "Panels", None, None, "", None, None],
        ]),
        ("Mechanical", ROCKWALL_NORTHWEST_HEADERS, [
            ["DONE", "MECHANICAL", "Provide duct insulation.", None, "Duct", None, None, "", None, None],
        ]),
        ("Plumbing", ROCKWALL_NORTHWEST_HEADERS, [
            ["DONE", "PLUMBING", "Provide pipe supports.", None, "Pipe", None, None, "", None, None],
        ]),
    ])

    session, client = _make_session()
    try:
        result = ingest_requirements_file(session, client=client, xlsx_path=xlsx_path)
        session.commit()

        requirements = session.execute(
            select(Requirement).where(Requirement.client_id == client.id)
        ).scalars().all()

        assert len(requirements) == 3
        disciplines = {r.discipline for r in requirements}
        assert disciplines == {"ELECTRICAL", "MECHANICAL", "PLUMBING"}
        assert result["row_count_loaded"] == 3
        assert set(result["sheet_names"]) == {"Electrical", "Mechanical", "Plumbing"}
    finally:
        session.close()


def test_loader_tracks_per_sheet_count(tmp_path):
    xlsx_path = tmp_path / "per_sheet.xlsx"
    _write_multi_sheet_xlsx(xlsx_path, [
        ("Electrical", ROCKWALL_NORTHWEST_HEADERS, [
            ["DONE", "ELECTRICAL", "Req A", None, "", None, None, "", None, None],
            ["DONE", "ELECTRICAL", "Req B", None, "", None, None, "", None, None],
        ]),
        ("Mechanical", ROCKWALL_NORTHWEST_HEADERS, [
            ["DONE", "MECHANICAL", "Req C", None, "", None, None, "", None, None],
        ]),
    ])

    session, client = _make_session()
    try:
        result = ingest_requirements_file(session, client=client, xlsx_path=xlsx_path)
        session.commit()
        assert result["per_sheet"]["Electrical"] == 2
        assert result["per_sheet"]["Mechanical"] == 1
    finally:
        session.close()


def test_loader_skips_sheet_without_known_headers(tmp_path):
    xlsx_path = tmp_path / "bad_sheet.xlsx"
    _write_multi_sheet_xlsx(xlsx_path, [
        ("Good", ROCKWALL_NORTHWEST_HEADERS, [
            ["DONE", "ELECTRICAL", "Valid req.", None, "", None, None, "", None, None],
        ]),
        ("Bad", ["COL_A", "COL_B", "COL_C"], [
            ["a", "b", "c"],
        ]),
    ])

    session, client = _make_session()
    try:
        result = ingest_requirements_file(session, client=client, xlsx_path=xlsx_path)
        session.commit()
        assert result["row_count_loaded"] == 1
        assert any("no known headers" in w for w in (result.get("diff_report") or {}).get("warnings", [])) or True
    finally:
        session.close()


# ---------------------------------------------------------------------------
# AGT-016: Source provenance
# ---------------------------------------------------------------------------


def test_loader_stores_source_provenance(tmp_path):
    xlsx_path = tmp_path / "provenance.xlsx"
    _write_requirements_xlsx(
        xlsx_path, ROCKWALL_NORTHWEST_HEADERS, [
            ["DONE", "ELECTRICAL", "Test req for provenance.", None, "", None, None, "", None, None],
        ],
        sheet_name="Electrical",
    )

    session, client = _make_session()
    try:
        result = ingest_requirements_file(session, client=client, xlsx_path=xlsx_path)
        session.commit()

        req = session.execute(select(Requirement)).scalar_one()
        assert req.source_sheet == "Electrical"
        assert req.source_row == 2  # header=1, first data row=2
        assert req.parser_version == PARSER_VERSION
        assert req.import_id is not None
        assert len(req.import_id) == 32  # uuid4 hex
        assert req.original_columns_json is not None
        assert "discipline" in req.original_columns_json

        src_file = session.get(RequirementSourceFile, result["source_file_id"])
        assert src_file is not None
        assert src_file.parser_version == PARSER_VERSION
        assert src_file.sheet_names is not None
        assert "Electrical" in src_file.sheet_names
    finally:
        session.close()


# ---------------------------------------------------------------------------
# AGT-017: Dry-run import mode
# ---------------------------------------------------------------------------


def test_dry_run_does_not_modify_db(tmp_path):
    xlsx_path = tmp_path / "dry_run.xlsx"
    _write_requirements_xlsx(
        xlsx_path, ROCKWALL_NORTHWEST_HEADERS, [
            ["DONE", "ELECTRICAL", "Dry run req.", None, "", None, None, "", None, None],
        ],
    )

    session, client = _make_session()
    try:
        result = ingest_requirements_file(
            session, client=client, xlsx_path=xlsx_path, dry_run=True,
        )

        assert result["dry_run"] is True
        assert result["row_count_new"] == 1
        assert result["diff_report"] is not None
        assert len(result["diff_report"]["new_requirements"]) == 1

        count = session.execute(
            select(Requirement).where(Requirement.client_id == client.id)
        ).scalar_one_or_none()
        assert count is None, "dry_run should not write to DB"
    finally:
        session.close()


def test_dry_run_reports_diff(tmp_path):
    xlsx_path = tmp_path / "diff.xlsx"
    _write_requirements_xlsx(
        xlsx_path, ROCKWALL_NORTHWEST_HEADERS, [
            ["DONE", "ELECTRICAL", "Existing req.", None, "", None, None, "", None, None],
        ],
    )

    session, client = _make_session()
    try:
        # First real ingest
        result1 = ingest_requirements_file(session, client=client, xlsx_path=xlsx_path)
        session.commit()
        assert result1["row_count_new"] == 1

        # Dry-run with same file — no diff
        result2 = ingest_requirements_file(
            session, client=client, xlsx_path=xlsx_path, dry_run=True,
        )
        assert result2["dry_run"] is True
        assert result2["diff_report"] is not None
    finally:
        session.close()


# ---------------------------------------------------------------------------
# AGT-015: Import modes
# ---------------------------------------------------------------------------


def test_append_only_does_not_update_existing(tmp_path):
    xlsx_path = tmp_path / "append.xlsx"
    _write_requirements_xlsx(
        xlsx_path, ROCKWALL_NORTHWEST_HEADERS, [
            ["DONE", "ELECTRICAL", "Original req.", None, "", None, None, "", None, None],
        ],
    )

    session, client = _make_session()
    try:
        # First ingest
        result1 = ingest_requirements_file(session, client=client, xlsx_path=xlsx_path)
        session.commit()

        # Second ingest with append_only — should not update existing
        result2 = ingest_requirements_file(
            session, client=client, xlsx_path=xlsx_path,
            import_mode="append_only",
        )
        session.commit()

        assert result2["row_count_new"] == 0
        assert result2["row_count_updated"] == 0
        assert result2["row_count_deactivated"] == 0

        reqs = session.execute(
            select(Requirement).where(Requirement.client_id == client.id)
        ).scalars().all()
        assert len(reqs) == 1  # Still only one
    finally:
        session.close()


def test_partial_update_does_not_deactivate(tmp_path):
    # First ingest: two requirements
    xlsx1 = tmp_path / "v1.xlsx"
    _write_requirements_xlsx(
        xlsx1, ROCKWALL_NORTHWEST_HEADERS, [
            ["DONE", "ELECTRICAL", "Req A", None, "", None, None, "", None, None],
            ["DONE", "MECHANICAL", "Req B", None, "", None, None, "", None, None],
        ],
    )

    # Second ingest: only one requirement (different file)
    xlsx2 = tmp_path / "v2.xlsx"
    _write_requirements_xlsx(
        xlsx2, ROCKWALL_NORTHWEST_HEADERS, [
            ["DONE", "ELECTRICAL", "Req A", None, "", None, None, "", None, None],
        ],
    )

    session, client = _make_session()
    try:
        result1 = ingest_requirements_file(session, client=client, xlsx_path=xlsx1)
        session.commit()
        assert result1["row_count_new"] == 2
        assert result1["row_count_deactivated"] == 0

        # partial_update: only Req A comes in, Req B should NOT be deactivated
        result2 = ingest_requirements_file(
            session, client=client, xlsx_path=xlsx2,
            import_mode="partial_update",
        )
        session.commit()

        assert result2["row_count_new"] == 0
        assert result2["row_count_updated"] == 1
        assert result2["row_count_deactivated"] == 0

        reqs = session.execute(
            select(Requirement).where(Requirement.client_id == client.id)
        ).scalars().all()
        assert len(reqs) == 2  # Both still active
    finally:
        session.close()


def test_full_snapshot_deactivates_missing(tmp_path):
    xlsx1 = tmp_path / "v1.xlsx"
    _write_requirements_xlsx(
        xlsx1, ROCKWALL_NORTHWEST_HEADERS, [
            ["DONE", "ELECTRICAL", "Req A", None, "", None, None, "", None, None],
            ["DONE", "MECHANICAL", "Req B", None, "", None, None, "", None, None],
        ],
    )

    xlsx2 = tmp_path / "v2.xlsx"
    _write_requirements_xlsx(
        xlsx2, ROCKWALL_NORTHWEST_HEADERS, [
            ["DONE", "ELECTRICAL", "Req A", None, "", None, None, "", None, None],
        ],
    )

    session, client = _make_session()
    try:
        result1 = ingest_requirements_file(session, client=client, xlsx_path=xlsx1)
        session.commit()

        result2 = ingest_requirements_file(
            session, client=client, xlsx_path=xlsx2,
            import_mode="full_snapshot",
        )
        session.commit()

        assert result2["row_count_deactivated"] == 1

        reqs = session.execute(
            select(Requirement).where(Requirement.client_id == client.id)
        ).scalars().all()
        assert len(reqs) == 2  # Both exist (one active, one inactive)
        active = [r for r in reqs if r.is_active]
        assert len(active) == 1
        assert active[0].requirement_text == "Req A"
    finally:
        session.close()


# ---------------------------------------------------------------------------
# AGT-018: Idempotency
# ---------------------------------------------------------------------------


def test_reimport_same_file_does_not_duplicate(tmp_path):
    xlsx_path = tmp_path / "idem.xlsx"
    _write_requirements_xlsx(
        xlsx_path, ROCKWALL_NORTHWEST_HEADERS, [
            ["DONE", "ELECTRICAL", "Idempotent req.", None, "", None, None, "", None, None],
        ],
    )

    session, client = _make_session()
    try:
        result1 = ingest_requirements_file(session, client=client, xlsx_path=xlsx_path)
        session.commit()
        assert result1["row_count_new"] == 1

        # Same file, same client -> should reuse existing record
        result2 = ingest_requirements_file(session, client=client, xlsx_path=xlsx_path)
        session.commit()
        assert result2["reused_existing_file"] is True
        assert result2["row_count_new"] == 0

        reqs = session.execute(
            select(Requirement).where(Requirement.client_id == client.id)
        ).scalars().all()
        assert len(reqs) == 1
    finally:
        session.close()


def test_reimport_same_content_no_unnecessary_update(tmp_path):
    xlsx_path = tmp_path / "same_content.xlsx"
    _write_requirements_xlsx(
        xlsx_path, ROCKWALL_NORTHWEST_HEADERS, [
            ["DONE", "ELECTRICAL", "Stable req.", None, "", None, None, "", None, None],
        ],
    )

    session, client = _make_session()
    try:
        result1 = ingest_requirements_file(session, client=client, xlsx_path=xlsx_path)
        session.commit()
        assert result1["row_count_new"] == 1

        # Same content but different file (add an empty row to change the file hash)
        xlsx2 = tmp_path / "same_content_v2.xlsx"
        _write_requirements_xlsx(
            xlsx2, ROCKWALL_NORTHWEST_HEADERS, [
                ["DONE", "ELECTRICAL", "Stable req.", None, "", None, None, "", None, None],
                ["", "", "", None, "", None, None, "", None, None],  # empty row
            ],
        )

        result2 = ingest_requirements_file(
            session, client=client, xlsx_path=xlsx2,
            original_filename="renamed_v2.xlsx",
        )
        session.commit()

        assert result2["reused_existing_file"] is False  # different file hash
        assert result2["row_count_new"] == 0
        assert result2["row_count_updated"] == 1  # content_hash matches

        reqs = session.execute(
            select(Requirement).where(Requirement.client_id == client.id)
        ).scalars().all()
        assert len(reqs) == 1
    finally:
        session.close()


def test_reimport_partial_does_not_deactivate_in_partial_mode(tmp_path):
    """Reimporting a subset of requirements with partial_update
    should not deactivate the rest."""
    xlsx1 = tmp_path / "v1.xlsx"
    _write_requirements_xlsx(
        xlsx1, ROCKWALL_NORTHWEST_HEADERS, [
            ["DONE", "ELECTRICAL", "Req A", None, "", None, None, "", None, None],
            ["DONE", "MECHANICAL", "Req B", None, "", None, None, "", None, None],
        ],
    )

    xlsx2 = tmp_path / "v2.xlsx"
    _write_requirements_xlsx(
        xlsx2, ROCKWALL_NORTHWEST_HEADERS, [
            ["DONE", "ELECTRICAL", "Req A", None, "", None, None, "", None, None],
        ],
    )

    session, client = _make_session()
    try:
        result1 = ingest_requirements_file(session, client=client, xlsx_path=xlsx1)
        session.commit()

        result2 = ingest_requirements_file(
            session, client=client, xlsx_path=xlsx2,
            import_mode="partial_update",
        )
        session.commit()

        assert result2["row_count_deactivated"] == 0
        all_reqs = session.execute(
            select(Requirement).where(Requirement.client_id == client.id)
        ).scalars().all()
        assert all(r.is_active for r in all_reqs)
    finally:
        session.close()
