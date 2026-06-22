"""Owner Requirements loader.

Ingests a SharePoint-style xlsx export of a district's requirement catalog into
the `client` / `requirement_source_file` / `requirement` tables.

Supports:
- Multi-sheet workbooks (all sheets are parsed)
- Header detection per sheet (first row with known headers wins)
- Provenance tracking (file, sheet, row, original columns)
- Import modes: full_snapshot, partial_update, append_only
- Dry-run mode (reports diff without writing to DB)
- Idempotent on (client_id, file_hash) and (client_id, content_hash)
"""
from __future__ import annotations

import hashlib
import json
import logging
import re
import uuid
from collections import Counter
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any, Literal

from openpyxl import load_workbook
from sqlalchemy import select, update
from sqlalchemy.orm import Session

from app.models import Client, Requirement, RequirementSourceFile

logger = logging.getLogger(__name__)

PARSER_VERSION = "2.0"

# ----------------------------------------------------------------------------
# Constants / normalization
# ----------------------------------------------------------------------------

CANONICAL_HEADERS = {
    "STATUS": "status",
    "DISCIPLINE": "discipline",
    "REQUIREMENT": "requirement",
    "LINKS": "links",
    "CATEGORY LIST": "category",
    "DATE UPDATED": "date_updated",
    "MODIFIED BY": "modified_by",
    "RESOURCE": "resource",
    "ITEM TYPE": "item_type",
    "PATH": "path",
}

VALID_DISCIPLINES = {
    "ELECTRICAL",
    "LIGHTING",
    "PLUMBING",
    "MECHANICAL",
    "TECHNOLOGY",
}

SENTINEL_PATTERNS = (
    re.compile(r"drag\s*&\s*drop\s+requirement\s+here", re.IGNORECASE),
    re.compile(r"\U0001f60e"),
)

MIN_DATE = datetime(1990, 1, 1, tzinfo=timezone.utc)

ImportMode = Literal["full_snapshot", "partial_update", "append_only"]

# ----------------------------------------------------------------------------
# Helpers
# ----------------------------------------------------------------------------


def _sha256_file(path: Path, chunk: int = 1024 * 1024) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        while True:
            buf = f.read(chunk)
            if not buf:
                break
            h.update(buf)
    return h.hexdigest()


def _sha256_text(*parts: str) -> str:
    return hashlib.sha256("|".join(parts).encode("utf-8")).hexdigest()


def _normalize_text(value: Any) -> str:
    if value is None:
        return ""
    s = str(value).replace("\r", " ").replace("\n", " | ")
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _is_sentinel(text: str) -> bool:
    return any(p.search(text) for p in SENTINEL_PATTERNS)


def _is_reference_only_requirement(req_text: str, links: str | None, resource: str | None) -> bool:
    if not req_text:
        return False
    normalized_text = req_text.lower().strip()
    reference_phrases = [
        "refer to links column",
        "refer to hyperlink",
        "hyperlink to",
        "refer to links",
    ]
    has_reference_phrase = any(phrase in normalized_text for phrase in reference_phrases)
    has_links_or_resource = bool(_normalize_text(links)) or bool(_normalize_text(resource))
    return has_reference_phrase and has_links_or_resource


def _normalize_discipline(value: Any) -> str:
    s = _normalize_text(value).upper()
    if not s:
        return "OTHER"
    if s in VALID_DISCIPLINES:
        return s
    if s.startswith("ELEC"):
        return "ELECTRICAL"
    if s.startswith("LIGHT"):
        return "LIGHTING"
    if s.startswith("PLUMB"):
        return "PLUMBING"
    if s.startswith("MECH"):
        return "MECHANICAL"
    if s.startswith("TECH"):
        return "TECHNOLOGY"
    return "OTHER"


def _normalize_status(value: Any) -> str | None:
    s = _normalize_text(value).upper()
    if not s:
        return None
    if "DONE" in s:
        return "DONE"
    if "NOT STARTED" in s or s == "NOT_STARTED":
        return "NOT_STARTED"
    return s


def _coerce_date(value: Any) -> datetime | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        dt = value
    elif isinstance(value, date):
        dt = datetime(value.year, value.month, value.day)
    else:
        try:
            dt = datetime.fromisoformat(str(value))
        except ValueError:
            return None
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    now = datetime.now(timezone.utc)
    if dt < MIN_DATE or dt.year > now.year + 5:
        return None
    return dt


_FILENAME_DATE_RE = re.compile(
    r"(?P<m>\d{1,2})[._\- ](?P<d>\d{1,2})[._\- ](?P<y>\d{2,4})"
)


def _parse_export_date_from_filename(name: str) -> date | None:
    m = _FILENAME_DATE_RE.search(name)
    if not m:
        return None
    try:
        y = int(m.group("y"))
        if y < 100:
            y += 2000
        return date(y, int(m.group("m")), int(m.group("d")))
    except ValueError:
        return None


def _detect_header_row(ws, min_row: int = 1, max_row: int = 10) -> int | None:
    """Score rows for header-like content. Returns the first row that has at
    least 3 known canonical headers, or None if none found."""
    for row_idx in range(min_row, min(max_row + 1, ws.max_row + 1)):
        matched = 0
        for col_idx in range(1, ws.max_column + 1):
            raw = ws.cell(row_idx, col_idx).value
            if raw is not None and str(raw).strip().upper() in CANONICAL_HEADERS:
                matched += 1
        if matched >= 3:
            return row_idx
    return None


def _build_header_map(ws, header_row: int) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        raw = ws.cell(header_row, col).value
        if raw is None:
            continue
        key = str(raw).strip().upper()
        if key in CANONICAL_HEADERS:
            mapping[CANONICAL_HEADERS[key]] = col
    return mapping


# ----------------------------------------------------------------------------
# Public API
# ----------------------------------------------------------------------------


def get_client_by_code(db: Session, code: str) -> Client | None:
    return db.execute(select(Client).where(Client.code == code)).scalar_one_or_none()


def ingest_requirements_file(
    db: Session,
    *,
    client: Client,
    xlsx_path: Path,
    original_filename: str | None = None,
    import_mode: ImportMode = "full_snapshot",
    dry_run: bool = False,
    parser_version: str = PARSER_VERSION,
) -> dict[str, Any]:
    """Ingest an xlsx file for the given client.

    Parameters
    ----------
    db : Session
    client : Client
    xlsx_path : Path
        Path to the .xlsx file on disk.
    original_filename : str, optional
        Override for display (defaults to xlsx_path.name).
    import_mode : str
        ``full_snapshot`` (default) — deactivates requirements not seen.
        ``partial_update`` — never deactivates; updates existing, adds new.
        ``append_only`` — only adds new requirements; never updates or deactivates.
    dry_run : bool
        If True, report the diff without writing to the database.
    parser_version : str
        Version string stored in provenance.

    Returns
    -------
    dict with counts, provenance, and optionally a diff_report (dry_run).
    """
    original_filename = original_filename or xlsx_path.name
    file_hash = _sha256_file(xlsx_path)
    import_id = uuid.uuid4().hex

    # Idempotency: same file bytes for same client.
    existing_source = db.execute(
        select(RequirementSourceFile).where(
            RequirementSourceFile.client_id == client.id,
            RequirementSourceFile.file_hash == file_hash,
        )
    ).scalar_one_or_none()

    reused_existing_file = False
    if existing_source is not None:
        logger.info(
            "Requirements file already ingested (client=%s hash=%s), reusing record",
            client.code,
            file_hash[:12],
        )
        reused_existing_file = True
        source_file = existing_source
    else:
        source_file = RequirementSourceFile(
            client_id=client.id,
            original_filename=original_filename,
            file_hash=file_hash,
            export_date=_parse_export_date_from_filename(original_filename),
            parser_version=parser_version,
        )

    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    sheet_names = list(wb.sheetnames)
    if not source_file.sheet_names:
        source_file.sheet_names = json.dumps(sheet_names)

    # Track per-sheet and per-discipline counts.
    per_discipline: Counter[str] = Counter()
    per_sheet: Counter[str] = Counter()

    # For full_snapshot: track all content hashes seen across all sheets.
    seen_hashes: set[str] = set()
    # Track which hashes are new vs updated vs unchanged.
    new_requirements: list[dict[str, Any]] = []
    updated_requirements: list[dict[str, Any]] = []
    unchanged_requirements: list[dict[str, Any]] = []
    warnings: list[str] = []
    errors: list[str] = []

    row_count_raw = 0
    row_count_skipped = 0
    row_count_new = 0
    row_count_updated = 0

    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        header_row = _detect_header_row(ws)
        if header_row is None:
            warnings.append(f"Sheet '{sheet_name}': no known headers found in first 10 rows, skipping")
            continue

        hmap = _build_header_map(ws, header_row)
        if "requirement" not in hmap or "discipline" not in hmap:
            warnings.append(
                f"Sheet '{sheet_name}': missing required columns 'requirement' and 'discipline'. "
                f"Found: {list(hmap.keys())}, skipping"
            )
            continue

        def _cell(row: tuple, key: str) -> Any:
            col = hmap.get(key)
            if col is None:
                return None
            idx = col - 1
            return row[idx] if 0 <= idx < len(row) else None

        sheet_row_count = 0

        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            row_count_raw += 1
            sheet_row_count += 1
            source_row_number = header_row + sheet_row_count

            req_text_raw = _cell(row, "requirement")
            req_text = _normalize_text(req_text_raw)

            if not req_text:
                row_count_skipped += 1
                continue
            if _is_sentinel(req_text):
                row_count_skipped += 1
                continue

            discipline = _normalize_discipline(_cell(row, "discipline"))
            content_hash = _sha256_text(discipline, req_text.lower())
            if content_hash in seen_hashes:
                row_count_skipped += 1
                continue
            seen_hashes.add(content_hash)

            category = _normalize_text(_cell(row, "category")) or None
            owner_status = _normalize_status(_cell(row, "status"))
            resource = _normalize_text(_cell(row, "resource")) or None
            links = _normalize_text(_cell(row, "links")) or None
            modified_by = _normalize_text(_cell(row, "modified_by")) or None
            date_updated = _coerce_date(_cell(row, "date_updated"))
            sharepoint_path = _normalize_text(_cell(row, "path")) or None

            original_columns: dict[str, str] = {}
            for canonical_name, col_idx in hmap.items():
                val = _cell(row, canonical_name)
                if val is not None:
                    original_columns[canonical_name] = str(val)

            existing_req = db.execute(
                select(Requirement).where(
                    Requirement.client_id == client.id,
                    Requirement.content_hash == content_hash,
                )
            ).scalar_one_or_none()

            is_actionable = not _is_reference_only_requirement(req_text, links, resource)

            provenance = {
                "source_sheet": sheet_name,
                "source_row": source_row_number,
                "original_columns_json": original_columns if original_columns else None,
                "parser_version": parser_version,
                "import_id": import_id,
            }

            if existing_req is None:
                if import_mode != "append_only":
                    row_data = dict(
                        client_id=client.id,
                        source_file_id=source_file.id,
                        discipline=discipline,
                        category=category,
                        requirement_text=req_text,
                        content_hash=content_hash,
                        owner_status=owner_status,
                        resource=resource,
                        links=links,
                        modified_by=modified_by,
                        date_updated=date_updated,
                        sharepoint_path=sharepoint_path,
                        is_actionable=is_actionable,
                        is_active=True,
                        **provenance,
                    )
                    new_requirements.append(row_data)
                    if not dry_run:
                        db.add(Requirement(**row_data))
                row_count_new += 1
            elif import_mode == "append_only":
                unchanged_requirements.append({
                    "content_hash": content_hash,
                    "requirement_text": req_text,
                    "discipline": discipline,
                    "reason": "append_only mode: existing requirement skipped",
                })
            else:
                if not dry_run:
                    existing_req.source_file_id = source_file.id
                    existing_req.source_sheet = sheet_name
                    existing_req.source_row = source_row_number
                    existing_req.original_columns_json = original_columns if original_columns else None
                    existing_req.parser_version = parser_version
                    existing_req.import_id = import_id
                    existing_req.category = category
                    existing_req.owner_status = owner_status
                    existing_req.resource = resource
                    existing_req.links = links
                    existing_req.modified_by = modified_by
                    existing_req.date_updated = date_updated
                    existing_req.sharepoint_path = sharepoint_path
                    existing_req.is_active = True
                    existing_req.is_actionable = is_actionable
                    existing_req.last_seen_at = datetime.now(timezone.utc)
                row_count_updated += 1
                updated_requirements.append({
                    "content_hash": content_hash,
                    "requirement_text": req_text,
                    "discipline": discipline,
                })

            per_discipline[discipline] += 1
            per_sheet[sheet_name] += 1

    # Apply deactivation based on import mode.
    deactivated_requirements: list[dict[str, Any]] = []

    if import_mode == "full_snapshot" and not dry_run and seen_hashes:
        existing_hashes = set(
            db.execute(
                select(Requirement.content_hash).where(
                    Requirement.client_id == client.id,
                    Requirement.is_active.is_(True),
                )
            ).scalars().all()
        )
        to_deactivate = existing_hashes - seen_hashes
        if to_deactivate:
            deactivated_rows = db.execute(
                select(Requirement).where(
                    Requirement.client_id == client.id,
                    Requirement.content_hash.in_(to_deactivate),
                )
            ).scalars().all()
            for r in deactivated_rows:
                deactivated_requirements.append({
                    "content_hash": r.content_hash,
                    "requirement_text": r.requirement_text,
                    "discipline": r.discipline,
                })
            db.execute(
                update(Requirement)
                .where(
                    Requirement.client_id == client.id,
                    Requirement.content_hash.in_(to_deactivate),
                )
                .values(is_active=False)
            )
    elif import_mode == "full_snapshot" and dry_run and seen_hashes:
        existing_hashes = set(
            db.execute(
                select(Requirement.content_hash).where(
                    Requirement.client_id == client.id,
                    Requirement.is_active.is_(True),
                )
            ).scalars().all()
        )
        to_deactivate = existing_hashes - seen_hashes
        if to_deactivate:
            rows = db.execute(
                select(Requirement).where(
                    Requirement.client_id == client.id,
                    Requirement.content_hash.in_(to_deactivate),
                )
            ).scalars().all()
            for r in rows:
                deactivated_requirements.append({
                    "content_hash": r.content_hash,
                    "requirement_text": r.requirement_text,
                    "discipline": r.discipline,
                })

    row_count_loaded = row_count_new + row_count_updated
    row_count_deactivated = len(deactivated_requirements)

    if not dry_run:
        if not reused_existing_file:
            db.add(source_file)
            db.flush()
        source_file.row_count_raw = row_count_raw
        source_file.row_count_loaded = row_count_loaded
        source_file.row_count_skipped = row_count_skipped
        db.flush()
    else:
        source_file.id = -1

    wb.close()

    diff_report = None
    if dry_run:
        diff_report = {
            "import_mode": import_mode,
            "new_requirements": new_requirements,
            "updated_requirements": updated_requirements,
            "deactivated_requirements": deactivated_requirements,
            "unchanged_requirements": unchanged_requirements,
            "warnings": warnings,
            "errors": errors,
            "per_discipline": dict(per_discipline),
            "per_sheet": dict(per_sheet),
        }

    return {
        "client_id": client.id,
        "source_file_id": source_file.id,
        "original_filename": original_filename,
        "file_hash": file_hash,
        "row_count_raw": row_count_raw,
        "row_count_loaded": row_count_loaded,
        "row_count_skipped": row_count_skipped,
        "row_count_new": row_count_new,
        "row_count_updated": row_count_updated,
        "row_count_deactivated": row_count_deactivated,
        "export_date": source_file.export_date,
        "reused_existing_file": reused_existing_file,
        "per_discipline": dict(per_discipline),
        "per_sheet": dict(per_sheet),
        "import_mode": import_mode,
        "import_id": import_id,
        "dry_run": dry_run,
        "diff_report": diff_report,
        "sheet_names": sheet_names,
        "parser_version": parser_version,
    }
